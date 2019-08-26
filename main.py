import asyncio
import itertools
from asyncio import sleep
from pathlib import Path
from urllib.parse import urljoin

from openpyxl import load_workbook
from pyppeteer import launch
from urllib3.util import parse_url

target_file_path = '404founder.xlsx'


def paginate(seq, page_size):
    """
    Iterates over the sequence in chunks of page_size maximum length.
    E.g.
    paginate([1,2,3,4,5,6,7,8,9,10], 3) --> 123 456 789 10
    :param seq: sequence to iterate
    :param page_size: max_size of the chunk

    """
    i = iter(seq)
    while True:
        page = tuple(itertools.islice(i, 0, page_size))
        if page:
            yield page
        else:
            return


def get_urls(path_to_file):
    wb = load_workbook(path_to_file)
    for sheet in wb.worksheets:

        values = iter(sheet.values)
        # skip headers
        next(values)
        for value in values:
            url = value[0]
            country = sheet.title
            if Path(f'images/{country}/{parse_url(url).hostname}.png').exists():
                continue

            yield country, url


def get_titles(path_to_file):
    wb = load_workbook(path_to_file)
    for sheet in wb.worksheets:
        yield sheet.title


async def do_screenshot(country, url):
    o = parse_url(url)
    wrong_url = urljoin(url, 'fuck')

    print(f'Starting new browser for {url}')
    browser = await launch(ignoreHTTPSErrors=True, autoClose=False)
    page = await browser.newPage()

    try:
        await page.goto(wrong_url, verify=False)
        await page.screenshot({'path': f'images/{country}/{o.hostname}.png', 'type': 'png'})
        print(f'Finished for country: {country}: {url}')
    finally:
        await sleep(1)
        await browser.close()


async def find404(path):
    for page in paginate(get_urls(path), 20):
        tasks = []
        for country, url in page:
            tasks.append(do_screenshot(country, url))

        results = await asyncio.gather(*tasks, return_exceptions=True)
        for r in filter(lambda x: x is not None, results):
            print(f"Error: {r}")
            pass


def create_folders(path):
    for country_name in get_titles(path):
        Path.mkdir(Path("images/{}".format(country_name)), parents=True, exist_ok=True)


create_folders(target_file_path)
asyncio.run(find404(target_file_path))
